import openpyxl
import os, string, sys
from lib import l

DIRS_SOCIUM = ['/media/da3/asteriskBeagleAl/Socium/2017/', '/media/da3/asteriskBeagleAl/Socium/2018/']

def isSNILS(snils):
    if snils != None:
        t = str(snils).replace('\n',' ').replace('  ', ' ').replace('  ', ' ').replace('  ', ' ').strip()
        if len(t) >= 11:
            if t[3] == '-' and t[7] == '-' and t[11] == ' ':
                return True
            else:
                return False
        else:
            return False
    return False

def isAudio(audio):
    if audio != None:
        t = str(audio).replace('\n',' ').replace('  ', ' ').replace('  ', ' ').replace('  ', ' ').strip()
        t1 = t.split('/')[len(t.split(('/'))) - 1]
        if len(t1) > 26:
            if t1[2] == '.' and t1[5] == '.' and t1[10] == '_' and t1[13] == '-' and t1[16] == '-':
                return ['длинный', t1]
            elif len(''.join([char for i, char in enumerate(t1) if char in string.digits and i < 26])) == 25 \
                    and t1[14] == '_':
                return ['короткий', t1]
            elif t.endswith('.mp3') or t.endswith('.wav'):
                return ['расширение', t1]
            else:
                return ['', audio]
        else:
            return ['', audio]
    return ['', audio]

def isSocium(audio):
    if audio != None:
        t = str(audio).replace('\n',' ').replace('  ', ' ').replace('  ', ' ').replace('  ', ' ').strip()
        t1 = t.split('/')[len(t.split(('/'))) - 1]
        if len(t1) > 26:
            if t1[2] == '.' and t1[5] == '.' and t1[10] == '_' and t1[13] == '-' and t1[16] == '-' \
                    and (t1[6:10] == 2017 or t1[6:10] == 2018):
                return True
            elif len(''.join([char for i, char in enumerate(t1) if char in string.digits and i < 26])) == 25 \
                    and t1[14] == '_' and (t1[:4] == 2017 or t1[:4] == 2018):
                return True
            else:
                return False
        else:
            return False
    return False


snils_audios = {}
snils_audios_fullpath = {}

for dir_socium in DIRS_SOCIUM:
    directories = os.listdir(dir_socium)
    for directory in directories:
        files = os.listdir(dir_socium + directory)
        for file in files:
            if file.endswith('.xlsx'):
                wb = openpyxl.load_workbook(filename=dir_socium + directory + '/'+ file, read_only=True)
                for sheetname in wb.sheetnames:
                    sheet = wb[sheetname]
                    if not sheet.max_row:
                        print('Файл', file, 'Excel некорректно сохранен OpenPyxl. Откройте и пересохраните его')
                        continue
                    print('\t накоплено связей СНИЛС-audio:', len(snils_audios_fullpath),'\n', dir_socium + directory +
                          '/'+ file + '!' + sheetname)
                    # В каждой строчке определяем где есть аудифайл и СНИЛС
                    table_j_end = 0  # Если больше 10 пустых ячеек - на следующую срочку
                    table_k_end = 0  # Если больше 10 пустых строчек - заканчиваем чтение таблицы
                    for j, row in enumerate(sheet.rows):
                        if table_j_end == 10 and j == 10:
                            break
                        snils = 0
                        audofiles = []
                        audofileExt = []
                        for k, cell in enumerate(row):
                            if cell.value != None:
                                table_j_end = 0
                                table_k_end = 0
                            else:
                                table_j_end += 1
                                table_k_end += 1
                            if table_k_end > 10:
                                break
                            if isSNILS(cell.value):
                                snils = l(cell.value)
                            else:
                                rezAudio = isAudio(cell.value)
                                if rezAudio[0]:
                                    if rezAudio[1].endswith('.wav') or rezAudio[1].endswith('.mp3'):
                                        rezAudioName = rezAudio[1][:-4]
                                    elif rezAudio[1].endswith('.') or rezAudio[1].endswith('/'):
                                        rezAudioName = rezAudio[1][:-1]
                                    else:
                                        rezAudioName = rezAudio[1]
                                    for audofileTek in files:
                                        if audofileTek[:-4] == rezAudioName:
                                            audofileExt.append(audofileTek)
                                            audofiles.append(rezAudioName)
                                            break
                        if snils and len(audofiles):
                            for i, audofile in enumerate(audofiles):
                                if snils_audios.get(snils, None):
                                    if audofile not in snils_audios[snils]:
                                        snils_audios[snils].append(audofile)
                                        snils_audios_fullpath[snils].append(dir_socium + directory + '/'+ audofileExt[i])
                                    else:
                                        print('\tДля СНИЛСа', snils, 'уже есть', dir_socium + directory + '/' + audofileExt[i])
                                else:
                                    snils_audios[snils] = [audofile]
                                    snils_audios_fullpath[snils] = [dir_socium + directory + '/'+ audofileExt[i]]
                        else:
                            if not snils and not len(audofiles):
                                pass
                            elif len(audofiles):
                                for audofile in audofiles:
                                    print('\tНе нашлось СНИЛСа для:', dir_socium + directory + '/' + audofileExt[i])
                            elif snils:
                                #print('\tВ директории', dir_socium + directory,'Не нашлось аудиофайла для СНИЛСа:', snils)
                                pass


