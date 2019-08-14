import openpyxl, traceback
import os, string, sys
from lib import l, fine_snils

DIRS_SOCIUM = ['/media/da3/asteriskBeagleAl/Socium/2017/', '/media/da3/asteriskBeagleAl/Socium/2018/']
FIND_CATALOG = '/media/da3/asteriskBeagleAl'
REESTRS = '/home/da3/Beagle/потеряшкиАудиозаписи/реестры/'

def isSNILS(snils):
    if snils != None:
        t = str(snils).replace('\n',' ').replace('  ', ' ').replace('  ', ' ').replace('  ', ' ').strip()
        if len(t) > 11:
            if t[3] == '-' and t[7] == '-' and t[11] == ' ':
                return True
            else:
                return False
        else:
            return False
    return False

def isAudio(audio):
    if audio != None:
        t = str(audio).replace('\n',' ').replace('  ', ' ').replace('  ', ' ').replace('  ', ' ').strip()
        t1 = t.split('/')[len(t.split(('/'))) - 1]
        if len(t1) > 26:
            if t1[2] == '.' and t1[5] == '.' and t1[10] == '_' and t1[13] == '-' and t1[16] == '-':
                return ['длинный', t1]
            elif len(''.join([char for i, char in enumerate(t1) if char in string.digits and i < 26])) == 25 \
                    and t1[14] == '_':
                return ['короткий', t1]
            elif t.endswith('.mp3') or t.endswith('.wav'):
                return ['расширение', t1]
            else:
                return ['', audio]
        else:
            return ['', audio]
    return ['', audio]

def isSocium(audio):
    if audio != None:
        t = str(audio).replace('\n',' ').replace('  ', ' ').replace('  ', ' ').replace('  ', ' ').strip()
        t1 = t.split('/')[len(t.split(('/'))) - 1]
        if len(t1) > 26:
            if t1[2] == '.' and t1[5] == '.' and t1[10] == '_' and t1[13] == '-' and t1[16] == '-' \
                    and (t1[6:10] == '2017' or t1[6:10] == '2018'):
                return True
            elif len(''.join([char for i, char in enumerate(t1) if char in string.digits and i < 26])) == 25 \
                    and t1[14] == '_' and (t1[:4] == '2017' or t1[:4] == '2018'):
                return True
            else:
                return False
        else:
            return False
    return False

# расшифровка любой ошибки
def full_tb_write(*args):
    if not args:
        exc_type, exc_val, exc_tb = sys.exc_info()
        traceback.print_tb(exc_tb, file=sys.stdout)
    elif len(args) == 3:
        exc_type, exc_val, exc_tb = args
        traceback.print_tb(exc_tb, file=sys.stdout)
    elif len(args) == 1:
        exc_type, exc_val, exc_tb = args[0].__class__, args[0], args[0].__traceback__
        traceback.print_tb(exc_tb, file=sys.stdout)

# Считываем рестры 2017 и 2018
files = os.listdir(REESTRS)
snilses = []
for file in files:
    if file.endswith('.xlsx'):
        wb = openpyxl.load_workbook(filename=REESTRS + file, read_only=True)
        for sheetname in wb.sheetnames:
            sheet = wb[sheetname]
            if not sheet.max_row:
                print('Файл', file, 'Excel некорректно сохранен OpenPyxl. Откройте и пересохраните его')
                continue
            table_j_end = 0  # Если больше 10 пустых ячеек - на следующую срочку
            table_k_end = 0  # Если больше 10 пустых строчек - заканчиваем чтение таблицы
            for j, row in enumerate(sheet.rows):
                if table_j_end == 10 and j == 10:
                    break
                snils = 0
                audiofiles = []
                for k, cell in enumerate(row):
                    if cell.value != None:
                        table_j_end = 0
                        table_k_end = 0
                    else:
                        table_j_end += 1
                        table_k_end += 1
                    if table_k_end > 10:
                        break
                    if isSNILS(cell.value):
                        snils = l(cell.value)
                        if snils not in snilses:
                            snilses.append(snils)

print('\n Уникальных СНИЛС в запросе:', len(snilses))

# Ищем все "надежные" файлы
snils_audios = {}
snils_audios_fullpath = {}
for dir_socium in DIRS_SOCIUM:
    directories = os.listdir(dir_socium)
    for directory in directories:
        files = os.listdir(dir_socium + directory)
        for file in files:
            if file.endswith('.xlsx'):
                wb = openpyxl.load_workbook(filename=dir_socium + directory + '/'+ file, read_only=True)
                for sheetname in wb.sheetnames:
                    sheet = wb[sheetname]
                    if not sheet.max_row:
                        print('Файл', file, 'Excel некорректно сохранен OpenPyxl. Откройте и пересохраните его')
                        continue
                    print('\t накоплено связей СНИЛС-audio:', len(snils_audios_fullpath),'\n', dir_socium + directory +
                          '/'+ file + '!' + sheetname)
                    # В каждой строчке определяем где есть аудифайл и СНИЛС
                    table_j_end = 0  # Если больше 10 пустых ячеек - на следующую срочку
                    table_k_end = 0  # Если больше 10 пустых строчек - заканчиваем чтение таблицы
                    for j, row in enumerate(sheet.rows):
                        if table_j_end == 10 and j == 10:
                            break
                        snils = 0
                        audiofiles = []
                        audiofileExt = []
                        for k, cell in enumerate(row):
                            if cell.value != None:
                                table_j_end = 0
                                table_k_end = 0
                            else:
                                table_j_end += 1
                                table_k_end += 1
                            if table_k_end > 10:
                                break
                            if isSNILS(cell.value):
                                snils = l(cell.value)
                            else:
                                rezAudio = isAudio(cell.value)
                                if rezAudio[0]:
                                    if rezAudio[1].endswith('.wav') or rezAudio[1].endswith('.mp3'):
                                        rezAudioName = rezAudio[1][:-4]
                                    elif rezAudio[1].endswith('.') or rezAudio[1].endswith('/'):
                                        rezAudioName = rezAudio[1][:-1]
                                    else:
                                        rezAudioName = rezAudio[1]
                                    for audiofileTek in files:
                                        if audiofileTek[:-4] == rezAudioName:
                                            audiofileExt.append(audiofileTek)
                                            audiofiles.append(rezAudioName)
                                            break
                        if snils and len(audiofiles):
                            for i, audiofile in enumerate(audiofiles):
                                if snils_audios.get(snils, None):
                                    if audiofile not in snils_audios[snils]:
                                        snils_audios[snils].append(audiofile)
                                        snils_audios_fullpath[snils].append(dir_socium + directory + '/'+
                                                                            audiofileExt[i])
                                    else:
                                        #print('\tДля СНИЛСа', snils, 'уже есть', dir_socium + directory + '/' +
                                        #      audiofileExt[i])
                                        pass
                                else:
                                    snils_audios[snils] = [audiofile]
                                    snils_audios_fullpath[snils] = [dir_socium + directory + '/'+ audiofileExt[i]]
                        else:
                            if not snils and not len(audiofiles):
                                pass
                            elif len(audiofiles):
                                for audiofile in audiofiles:
                                    print('\tНе нашлось СНИЛСа для:', dir_socium + directory + '/' + audiofileExt[i])
                            elif snils:
                                #print('\tВ директории', dir_socium + directory,'Не нашлось аудиофайла для СНИЛСа:', snils)
                                pass

# Реестр из надежных источников
wb = openpyxl.Workbook(write_only=True)
ws = wb.create_sheet('Надежные')
for snils in snilses:
    if snils_audios_fullpath.get(snils, None):
        ws.append([fine_snils(snils)] + snils_audios_fullpath[snils])
wb.save('Надежные.xlsx')

# ищет все файлы с именем filename во всех подкаталогах каталога catalog
all_audiofiles = []
for root, dirs, files in os.walk(FIND_CATALOG):
    all_audiofiles += [os.path.join(root, name) for name in files if isSocium(name)]
all_audiofilesExt = {}
for all_audiofile in all_audiofiles:
    rezAudioName = all_audiofile.split('/')[len(all_audiofile.split(('/'))) - 1]
    if rezAudioName.endswith('.wav') or rezAudioName.endswith('.mp3'):
        rezAudioName = rezAudioName[:-4]
    if all_audiofilesExt.get(rezAudioName, None):
        if all_audiofile not in all_audiofilesExt[rezAudioName]:
            all_audiofilesExt[rezAudioName].append(all_audiofile)
    else:
        all_audiofilesExt[rezAudioName] = [all_audiofile]

all_xlsxfiles = []
for root, dirs, files in os.walk(FIND_CATALOG):
    all_xlsxfiles += [os.path.join(root, name) for name in files if name.endswith('.xlsx')]
all_snils_audios = {}
max_all_xlsxfiles = len(all_xlsxfiles)
for jj in range(0, max_all_xlsxfiles):
    try:
        all_xlsxfile = all_xlsxfiles[jj]
        wb = openpyxl.load_workbook(filename=all_xlsxfile, read_only=True)
        for sheetname in wb.sheetnames:
            sheet = wb[sheetname]
            if not sheet.max_row:
                print('Файл', all_xlsxfile, 'Excel некорректно сохранен OpenPyxl. Откройте и пересохраните его')
                continue
            print('\t накоплено связей СНИЛС-audio:', len(all_snils_audios), '\n', all_xlsxfile + '!' + sheetname + ' ('
                  + str(jj) + ' из ' + str(max_all_xlsxfiles))
            # Аудиофайл[СНИЛС]
            table_j_end = 0  # Если больше 10 пустых ячеек - на следующую срочку
            table_k_end = 0  # Если больше 10 пустых строчек - заканчиваем чтение таблицы
            for j, row in enumerate(sheet.rows):
                if table_j_end == 10 and j == 10:
                    break
                snils = 0
                audiofiles = []
                for k, cell in enumerate(row):
                    if cell.value != None:
                        table_j_end = 0
                        table_k_end = 0
                    else:
                        table_j_end += 1
                        table_k_end += 1
                    if table_k_end > 10:
                        break
                    if isSNILS(cell.value):
                        snils = l(cell.value)
                    else:
                        rezAudio = isAudio(cell.value)
                        if rezAudio[0]:
                            if rezAudio[1].endswith('.wav') or rezAudio[1].endswith('.mp3'):
                                rezAudioName = rezAudio[1][:-4]
                            elif rezAudio[1].endswith('.') or rezAudio[1].endswith('/'):
                                rezAudioName = rezAudio[1][:-1]
                            else:
                                rezAudioName = rezAudio[1]
                            audiofiles.append(rezAudioName)
                if snils and len(audiofiles):
                    for i, audiofile in enumerate(audiofiles):
                        if all_snils_audios.get(snils, None):
                            if audiofile not in all_snils_audios[snils]:
                                all_snils_audios[snils].append(audiofile)
                        else:
                            all_snils_audios[snils] = [audiofile]
                else:
                    if not snils and not len(audiofiles):
                        pass
                    elif len(audiofiles):
                        for audiofile in audiofiles:
                            rezAudio = isAudio(audiofile)
                            if rezAudio[0] == 'длинный':
                                snils = int(rezAudio[1][20:31])
                                if all_snils_audios.get(snils, None):
                                    if audiofile not in all_snils_audios[snils]:
                                        all_snils_audios[snils].append(audiofile)
                                else:
                                    all_snils_audios[snils] = [audiofile]
                            else:
                                print('\tНе нашлось СНИЛСа для:', audiofile)
                    elif snils:
                        # print('\tВ директории', dir_socium + directory,'Не нашлось аудиофайла для СНИЛСа:', snils)
                        pass
    except Exception as e:
        full_tb_write(e)

# Собираем сначала из точных источников, потом из второстепенных
wb = openpyxl.Workbook(write_only=True)
ws = wb.create_sheet('Остальные')
shure_rez = {}
problem_rez = {}
for snils in snilses:
    if snils_audios_fullpath.get(snils, None):
        shure_rez[snils] = snils_audios_fullpath[snils]
    elif all_snils_audios.get(snils):
        all_snils_audio_vars = []                       # Собираем все файлы сюда
        for all_snils_audio in all_snils_audios:
            if all_audiofilesExt.get(all_snils_audio):
                for all_audiofileExt in all_audiofilesExt[all_snils_audio]:
                    all_snils_audio_vars.append(all_audiofileExt)
        problem_rez[snils] = all_snils_audio_vars
        ws.append([fine_snils(snils)] + all_snils_audio_vars)
wb.save('Остальные.xlsx')
print('Закрыто СНИЛСов из точных источников:', len(shure_rez))
print('Закрыто СНИЛСов из остальных источников:', len(problem_rez))








