import openpyxl, traceback
from mysql.connector import MySQLConnection
import os, string, sys, shutil
from collections import Counter
from lib import l, fine_snils, fine_snils_, read_config

OUTPUT_CATALOG = '/media/da3/backup/'
DIRS_SOCIUM = ['/media/da3/asteriskBeagleAl/Socium/2017/', '/media/da3/asteriskBeagleAl/Socium/2018/']
FIND_CATALOG = '/media/da3/asteriskBeagleAl'
TRUSTREESTR = '/home/da3/Beagle/потеряшкиАудиозаписи/Надежные.xlsx'
PROBLEMREESTR = '/home/da3/Beagle/потеряшкиАудиозаписи/Остальные.xlsx'
REESTRS = '/home/da3/Beagle/потеряшкиАудиозаписи/реестры/'

def isSNILS(snils):
    if snils != None:
        t = str(snils).replace('\n',' ').replace('  ', ' ').replace('  ', ' ').replace('  ', ' ').strip()
        if len(t) > 11:
            if t[3] == '-' and t[7] == '-' and (t[11] == ' ' or t[11] == '_'):
                return True
            else:
                return False
        else:
            return False
    return False

def isAudio(audio):
    if audio != None:
        t = str(audio).replace('\n',' ').replace('  ', ' ').replace('  ', ' ').replace('  ', ' ').strip()
        t1 = t.split('/')[len(t.split(('/'))) - 1]
        if t1.endswith('.'):
            t1 = t1[:-1]
        if t1.endswith('.mp3') or t1.endswith('.wav'):
            t1 = t1[:-4]
        if len(t1) > 26:
            if t1[2] == '.' and t1[5] == '.' and t1[10] == '_' and (t1[13] == '-' or t1[13] == '_') and \
                    (t1[16] == '-' or t1[16] == '_'):
                return ['длинный', t1]
            elif len(''.join([char for i, char in enumerate(t1) if char in string.digits and i < 26])) == 25 \
                    and t1[14] == '_':
                return ['короткий', t1]
            else:
                return ['', audio]
        else:
            return ['', audio]
    return ['', audio]

def isSocium(audio):
    if audio != None:
        t = str(audio).replace('\n',' ').replace('  ', ' ').replace('  ', ' ').replace('  ', ' ').strip()
        t1 = t.split('/')[len(t.split(('/'))) - 1]
        if len(t1) > 26:
            if t1[2] == '.' and t1[5] == '.' and t1[10] == '_' and (t1[13] == '-' or t1[13] == '_') and \
                    (t1[16] == '-' or t1[16] == '_') and (t1[6:10] == '2017' or t1[6:10] == '2018'):
                return True
            elif len(''.join([char for i, char in enumerate(t1) if char in string.digits and i < 26])) == 25 \
                    and t1[14] == '_' and (t1[:4] == '2017' or t1[:4] == '2018'):
                return True
            else:
                return False
        else:
            return False
    return False

# расшифровка любой ошибки
def full_tb_write(*args):
    if not args:
        exc_type, exc_val, exc_tb = sys.exc_info()
        traceback.print_tb(exc_tb, file=sys.stdout)
    elif len(args) == 3:
        exc_type, exc_val, exc_tb = args
        traceback.print_tb(exc_tb, file=sys.stdout)
    elif len(args) == 1:
        exc_type, exc_val, exc_tb = args[0].__class__, args[0], args[0].__traceback__
        traceback.print_tb(exc_tb, file=sys.stdout)

# Ищем все аудиофайлы isSocium() == True во всех подкаталогах каталога FIND_CATALOG
all_audiofiles = []
walking = list(os.walk(FIND_CATALOG))
for root, dirs, files in walking:
    all_audiofiles += [os.path.join(root, name) for name in files if isSocium(name)]
all_audiofilesExt = {}
for all_audiofile in all_audiofiles:
    rezAudioName = all_audiofile.split('/')[len(all_audiofile.split(('/'))) - 1]
    if rezAudioName.endswith('.wav') or rezAudioName.endswith('.mp3'):
        rezAudioName = rezAudioName[:-4]
    if all_audiofilesExt.get(rezAudioName, None):
        if all_audiofile not in all_audiofilesExt[rezAudioName]:
            all_audiofilesExt[rezAudioName].append(all_audiofile)
    else:
        all_audiofilesExt[rezAudioName] = [all_audiofile]

# Ищем ссылки на аудиофайлы во всех .xlsx файлах
snilsesProblem = {}
snilsesProblemShort = {}
all_xlsxfiles = []
for root, dirs, files in os.walk(FIND_CATALOG):
    all_xlsxfiles += [os.path.join(root, name) for name in files if name.endswith('.xlsx')]
all_snils_audios = {}
max_all_xlsxfiles = len(all_xlsxfiles)
for jj in range(0, max_all_xlsxfiles):
    try:
        all_xlsxfile = all_xlsxfiles[jj]
        wb = openpyxl.load_workbook(filename=all_xlsxfile, read_only=True)
        for sheetname in wb.sheetnames:
            sheet = wb[sheetname]
            if not sheet.max_row:
                print('Файл', all_xlsxfile, 'Excel некорректно сохранен OpenPyxl. Откройте и пересохраните его')
                continue
            print('\t накоплено связей СНИЛС-audio:', len(snilsesProblem), '\n', all_xlsxfile + '!' + sheetname + ' ('
                  + str(jj) + ' из ' + str(max_all_xlsxfiles) + ')')
            # Аудиофайл[СНИЛС]
            table_j_end = 0  # Если больше 10 пустых ячеек - на следующую срочку
            table_k_end = 0  # Если больше 10 пустых строчек - заканчиваем чтение таблицы
            for j, row in enumerate(sheet.rows):
                if table_j_end == 10 and j == 10:
                    break
                snils = 0
                audiofiles = []
                for k, cell in enumerate(row):
                    if cell.value != None:
                        table_j_end = 0
                        table_k_end = 0
                    else:
                        table_j_end += 1
                        table_k_end += 1
                    if table_k_end > 10:
                        break
                    if isSNILS(cell.value):
                        snils = l(cell.value)
                    else:
                        rezAudio = isAudio(cell.value)
                        if rezAudio[0]:
                            if rezAudio[1].endswith('.wav') or rezAudio[1].endswith('.mp3'):
                                rezAudioName = rezAudio[1][:-4]
                            elif rezAudio[1].endswith('.') or rezAudio[1].endswith('/'):
                                rezAudioName = rezAudio[1][:-1]
                            else:
                                rezAudioName = rezAudio[1]
                            audiofiles.append(rezAudioName)
                if snils and len(audiofiles):
                    for audiofile in audiofiles:
                        if all_audiofilesExt.get(audiofile, None):
                            if snilsesProblem.get(snils, None):
                                for temp in all_audiofilesExt[audiofile]:
                                    if temp not in snilsesProblem[snils]:
                                        snilsesProblem[snils].append(temp)
                                if audiofile not in snilsesProblemShort[snils]:
                                    snilsesProblemShort[snils].append(audiofile)
                            else:
                                snilsesProblem[snils] = all_audiofilesExt[audiofile]
                                snilsesProblemShort[snils] = [audiofile]
    except Exception as e:
        full_tb_write(e)

# Вытаскиваем словарь phonesSNILSES[телефон]=[СНИЛС1,...,СНИЛСn] из Сатурна для Социума
dbconfig_crm = read_config(filename='asocium.ini', section='crm')
dbconn = MySQLConnection(**dbconfig_crm)
cursor = dbconn.cursor()
sql = 'SELECT phone, snils FROM lekarh.alone_phone_snils'
cursor.execute(sql)
phonesSNILSES = {}
rows = cursor.fetchall()
for row in rows:
    if phonesSNILSES.get(row[0], None):
        if row[1] not in phonesSNILSES[row[0]]:
            phonesSNILSES[row[0]].append(row[1])
    else:
        phonesSNILSES[row[0]] = [row[1]]

# Считываем рестры 2017 и 2018
files = os.listdir(REESTRS)
snilsesI = []
for file in files:
    if file.endswith('.xlsx'):
        wb = openpyxl.load_workbook(filename=REESTRS + file, read_only=True)
        for sheetname in wb.sheetnames:
            sheet = wb[sheetname]
            if not sheet.max_row:
                print('Файл', file, 'Excel некорректно сохранен OpenPyxl. Откройте и пересохраните его')
                continue
            table_j_end = 0  # Если больше 10 пустых ячеек - на следующую срочку
            table_k_end = 0  # Если больше 10 пустых строчек - заканчиваем чтение таблицы
            for j, row in enumerate(sheet.rows):
                if table_j_end == 10 and j == 10:
                    break
                snils = 0
                audiofiles = []
                for k, cell in enumerate(row):
                    if cell.value != None:
                        table_j_end = 0
                        table_k_end = 0
                    else:
                        table_j_end += 1
                        table_k_end += 1
                    if table_k_end > 10:
                        break
                    if isSNILS(cell.value):
                        snils = l(cell.value)
                        if snils not in snilsesI:
                            snilsesI.append(snils)
snilsesInput = tuple(snilsesI)
print('\n Уникальных СНИЛС в запросе:', len(snilsesInput))

# Ищем все "надежные" файлы
snils_audios = {}
snils_audios_fullpath = {}
for dir_socium in DIRS_SOCIUM:
    directories = os.listdir(dir_socium)
    for directory in directories:
        files = os.listdir(dir_socium + directory)
        for file in files:
            if file.endswith('.xlsx'):
                wb = openpyxl.load_workbook(filename=dir_socium + directory + '/'+ file, read_only=True)
                for sheetname in wb.sheetnames:
                    sheet = wb[sheetname]
                    if not sheet.max_row:
                        print('Файл', file, 'Excel некорректно сохранен OpenPyxl. Откройте и пересохраните его')
                        continue
                    print('\t накоплено связей СНИЛС-audio:', len(snils_audios_fullpath),'\n', dir_socium + directory +
                          '/'+ file + '!' + sheetname)
                    # В каждой строчке определяем где есть аудифайл и СНИЛС
                    table_j_end = 0  # Если больше 10 пустых ячеек - на следующую срочку
                    table_k_end = 0  # Если больше 10 пустых строчек - заканчиваем чтение таблицы
                    for j, row in enumerate(sheet.rows):
                        if table_j_end == 10 and j == 10:
                            break
                        snils = 0
                        audiofiles = []
                        audiofileExt = []
                        for k, cell in enumerate(row):
                            if cell.value != None:
                                table_j_end = 0
                                table_k_end = 0
                            else:
                                table_j_end += 1
                                table_k_end += 1
                            if table_k_end > 10:
                                break
                            if isSNILS(cell.value):
                                snils = l(cell.value)
                            else:
                                rezAudio = isAudio(cell.value)
                                if rezAudio[0]:
                                    if rezAudio[1].endswith('.wav') or rezAudio[1].endswith('.mp3'):
                                        rezAudioName = rezAudio[1][:-4]
                                    elif rezAudio[1].endswith('.') or rezAudio[1].endswith('/'):
                                        rezAudioName = rezAudio[1][:-1]
                                    else:
                                        rezAudioName = rezAudio[1]
                                    for audiofileTek in files:
                                        if audiofileTek[:-4] == rezAudioName:
                                            audiofileExt.append(audiofileTek)
                                            audiofiles.append(rezAudioName)
                                            break
                        if snils and len(audiofiles):
                            for i, audiofile in enumerate(audiofiles):
                                if snils_audios.get(snils, None):
                                    if audiofile not in snils_audios[snils]:
                                        snils_audios[snils].append(audiofile)
                                        snils_audios_fullpath[snils].append(dir_socium + directory + '/'+
                                                                            audiofileExt[i])
                                    else:
                                        #print('\tДля СНИЛСа', snils, 'уже есть', dir_socium + directory + '/' +
                                        #      audiofileExt[i])
                                        pass
                                else:
                                    snils_audios[snils] = [audiofile]
                                    snils_audios_fullpath[snils] = [dir_socium + directory + '/'+ audiofileExt[i]]
                        else:
                            if not snils and not len(audiofiles):
                                pass
                            elif len(audiofiles):
                                for audiofile in audiofiles:
                                    print('\tНе нашлось СНИЛСа для:', dir_socium + directory + '/' + audiofileExt[i])
                            elif snils:
                                #print('\tВ директории', dir_socium + directory,'Не нашлось аудиофайла для СНИЛСа:', snils)
                                pass

# Реестр из надежных источников
snilsesTrust = {}
snilsesTrustShort = {}
for j, snils in enumerate(snils_audios_fullpath):
    snilsTrustAudios = []
    for audio in snils_audios_fullpath[snils]:
        snilsTrustAudio = isAudio(audio)
        if snilsTrustAudio[1] not in snilsTrustAudios:
            snilsTrustAudios.append(snilsTrustAudio[1])
        if snilsesTrust.get(snils, None):
            if audio not in snilsesTrust[snils]:
                snilsesTrust[snils].append(audio)
        else:
            snilsesTrust[snils] = [audio]
    snilsesTrustShort[snils] = snilsTrustAudios
snilsesTrustTuple = tuple(snilsesTrust.keys())
print('\n Уже найдено СНИЛС из надежных источников:', len(snilsesTrustTuple))

# Перебираем файлы и дозаполняем массив [СНИЛС] = [файл1, файл2...]
for all_audiofileExt in all_audiofilesExt:
    all_audiofile = isAudio(all_audiofileExt)
    snilsesTek = []
    if all_audiofile[0] == 'длинный':
        snilsesTek = [l(all_audiofile[1][20:31])]
    elif all_audiofile[0] == 'короткий':
        phone = l(all_audiofile[1][15:26])
        # через телефон найти СНИЛС (если есть)
        if phonesSNILSES.get(phone, None):
            for snils in phonesSNILSES[phone]:
                snilsesTek.append(snils)
    for snils in snilsesTek:
        if snilsesProblem.get(snils, None):
            for temp in all_audiofilesExt[all_audiofileExt]:
                if temp not in snilsesProblem[snils]:
                    snilsesProblem[snils].append(temp)
            if all_audiofile[1] not in snilsesProblemShort[snils]:
                snilsesProblemShort[snils].append(all_audiofile[1])
        else:
            snilsesProblem[snils] = all_audiofilesExt[all_audiofileExt]
            snilsesProblemShort[snils] = [all_audiofile[1]]

# Сохраняем реестр из надежных источников
wb = openpyxl.Workbook(write_only=True)
ws = wb.create_sheet('Надежные')
for snils in snilsesTrustTuple:
    if snils_audios_fullpath.get(snils, None):
        ws.append([fine_snils(snils)] + snils_audios_fullpath[snils])
wb.save('Надежные.xlsx')

# Сохраняем реестр из второстепенных источников
wb = openpyxl.Workbook(write_only=True)
ws = wb.create_sheet('Остальные')
problem_count = 0
for snils in snilsesInput:
    if snils not in snilsesTrustTuple:
        if snilsesProblem.get(snils, None):
            ws.append([fine_snils(snils)] + snilsesProblem[snils])
            problem_count += 1
wb.save('Остальные.xlsx')
print('Закрыто СНИЛСов из точных источников:', len(snilsesTrust))
print('Закрыто СНИЛСов из остальных источников:', problem_count)

# Разбираем надежные аудиофайлы по папочкам
for i, snils in enumerate(snilsesTrust):
    sucess = False
    while not sucess:
        try:
            if not os.path.exists(OUTPUT_CATALOG + 'Выгрузки/' + fine_snils_(snils)):
                os.mkdir(OUTPUT_CATALOG + 'Выгрузки/' + fine_snils_(snils))
            audiofilesShort = []
            for audiofile in snilsesTrust[snils]:
                audiofileShort = isAudio(audiofile)[1]
                if os.path.exists(audiofile): #.replace(FIND_CATALOG, CHANGE_ON_WINDOWS)):
                    if audiofileShort in audiofilesShort:
                        if not os.path.exists(OUTPUT_CATALOG + 'Выгрузки/' + fine_snils_(snils) + '/' + audiofileShort +
                                              '-' + str(Counter(audiofilesShort)[audiofileShort]) + audiofile[-4:]):
                            shutil.copy(audiofile #.replace(FIND_CATALOG, CHANGE_ON_WINDOWS)
                                        , OUTPUT_CATALOG + 'Выгрузки/' + fine_snils_(snils) + '/' + audiofileShort +
                                        '-' + str(Counter(audiofilesShort)[audiofileShort]) + audiofile[-4:])
                            audiofilesShort.append(audiofileShort)
                    else:
                        if not os.path.exists(OUTPUT_CATALOG + 'Выгрузки/' + fine_snils_(snils) + '/' +
                                              audiofileShort + audiofile[-4:]):
                            shutil.copy(audiofile #.replace(FIND_CATALOG, CHANGE_ON_WINDOWS)
                                        , OUTPUT_CATALOG + 'Выгрузки/' + fine_snils_(snils) + '/' + audiofileShort +
                                        audiofile[-4:])
                            audiofilesShort.append(audiofileShort)
                else:
                    print('!!! Нет исходного файла', audiofile)
            sucess = True
        except Exception as e:
            full_tb_write(e)
            print('Ошибка - пробуем ещё раз')
    print('Скопировано', i, 'из', len(snilsesTrust))

print('\nТеперь Остальные\n')

# Разбираем остальные аудиофайлы по папочкам
for i, snils in enumerate(snilsesProblem):
    sucess = False
    while not sucess:
        try:
            if not os.path.exists(OUTPUT_CATALOG + 'Остальные/' + fine_snils_(snils)):
                os.mkdir(OUTPUT_CATALOG + 'Остальные/' + fine_snils_(snils))
            audiofilesShort = []
            for audiofile in snilsesProblem[snils]:
                audiofileShort = isAudio(audiofile)[1]
                if os.path.exists(audiofile): #.replace(FIND_CATALOG, CHANGE_ON_WINDOWS)):
                    if audiofileShort in audiofilesShort:
                        if not os.path.exists(OUTPUT_CATALOG + 'Остальные/' + fine_snils_(snils) + '/' + audiofileShort +
                                              '-' + str(Counter(audiofilesShort)[audiofileShort]) + audiofile[-4:]):
                            shutil.copy(audiofile #.replace(FIND_CATALOG, CHANGE_ON_WINDOWS)
                                        , OUTPUT_CATALOG + 'Остальные/' + fine_snils_(snils) + '/' + audiofileShort +
                                        '-' + str(Counter(audiofilesShort)[audiofileShort]) + audiofile[-4:])
                            audiofilesShort.append(audiofileShort)
                    else:
                        if not os.path.exists(OUTPUT_CATALOG + 'Остальные/' + fine_snils_(snils) + '/' +
                                              audiofileShort + audiofile[-4:]):
                            shutil.copy(audiofile #.replace(FIND_CATALOG, CHANGE_ON_WINDOWS)
                                        , OUTPUT_CATALOG + 'Остальные/' + fine_snils_(snils) + '/' + audiofileShort +
                                        audiofile[-4:])
                            audiofilesShort.append(audiofileShort)
                else:
                    print('!!! Нет исходного файла', audiofile)
            sucess = True
        except Exception as e:
            full_tb_write(e)
            print('Ошибка - пробуем ещё раз')
    print('Скопировано', i, 'из', len(snilsesProblem))








