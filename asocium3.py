import openpyxl, traceback
from mysql.connector import MySQLConnection
import os, string, sys
from lib import l, fine_snils, read_config

DIRS_SOCIUM = ['/media/da3/asteriskBeagleAl/Socium/2017/', '/media/da3/asteriskBeagleAl/Socium/2018/']
FIND_CATALOG = '/media/da3/asteriskBeagleAl'
TRUSTREESTR = '/home/da3/Beagle/потеряшкиАудиозаписи/Надежные.xlsx'
PROBLEMREESTR = '/home/da3/Beagle/потеряшкиАудиозаписи/Остальные.xlsx'
REESTRS = '/home/da3/Beagle/потеряшкиАудиозаписи/реестры/'

def isSNILS(snils):
    if snils != None:
        t = str(snils).replace('\n',' ').replace('  ', ' ').replace('  ', ' ').replace('  ', ' ').strip()
        if len(t) > 11:
            if t[3] == '-' and t[7] == '-' and t[11] == ' ':
                return True
            else:
                return False
        else:
            return False
    return False

def isAudio(audio):
    if audio != None:
        t = str(audio).replace('\n',' ').replace('  ', ' ').replace('  ', ' ').replace('  ', ' ').strip()
        t1 = t.split('/')[len(t.split(('/'))) - 1]
        if t1.endswith('.'):
            t1 = t1[:-1]
        if t1.endswith('.mp3') or t1.endswith('.wav'):
            t1 = t1[:-4]
        if len(t1) > 26:
            if t1[2] == '.' and t1[5] == '.' and t1[10] == '_' and t1[13] == '-' and t1[16] == '-':
                return ['длинный', t1]
            elif len(''.join([char for i, char in enumerate(t1) if char in string.digits and i < 26])) == 25 \
                    and t1[14] == '_':
                return ['короткий', t1]
            else:
                return ['', audio]
        else:
            return ['', audio]
    return ['', audio]

def isSocium(audio):
    if audio != None:
        t = str(audio).replace('\n',' ').replace('  ', ' ').replace('  ', ' ').replace('  ', ' ').strip()
        t1 = t.split('/')[len(t.split(('/'))) - 1]
        if len(t1) > 26:
            if t1[2] == '.' and t1[5] == '.' and t1[10] == '_' and t1[13] == '-' and t1[16] == '-' \
                    and (t1[6:10] == '2017' or t1[6:10] == '2018'):
                return True
            elif len(''.join([char for i, char in enumerate(t1) if char in string.digits and i < 26])) == 25 \
                    and t1[14] == '_' and (t1[:4] == '2017' or t1[:4] == '2018'):
                return True
            else:
                return False
        else:
            return False
    return False

# расшифровка любой ошибки
def full_tb_write(*args):
    if not args:
        exc_type, exc_val, exc_tb = sys.exc_info()
        traceback.print_tb(exc_tb, file=sys.stdout)
    elif len(args) == 3:
        exc_type, exc_val, exc_tb = args
        traceback.print_tb(exc_tb, file=sys.stdout)
    elif len(args) == 1:
        exc_type, exc_val, exc_tb = args[0].__class__, args[0], args[0].__traceback__
        traceback.print_tb(exc_tb, file=sys.stdout)

# Считываем рестр надежных найденных за 2017 и 2018
snilsesT = []
wb = openpyxl.load_workbook(filename=TRUSTREESTR, read_only=True)
for sheetname in wb.sheetnames:
    sheet = wb[sheetname]
    if not sheet.max_row:
        print('Файл', TRUSTREESTR, 'Excel некорректно сохранен OpenPyxl. Откройте и пересохраните его')
        continue
    for j, row in enumerate(sheet.rows):
        for k, cell in enumerate(row):
            if k == 0:
                snilsesT.append(l(cell.value))
snilsesTrust = tuple(snilsesT)
print('\n Уже найдено СНИЛС из надежных источников:', len(snilsesTrust))

# Считываем рестры 2017 и 2018
files = os.listdir(REESTRS)
snilsesI = []
for file in files:
    if file.endswith('.xlsx'):
        wb = openpyxl.load_workbook(filename=REESTRS + file, read_only=True)
        for sheetname in wb.sheetnames:
            sheet = wb[sheetname]
            if not sheet.max_row:
                print('Файл', file, 'Excel некорректно сохранен OpenPyxl. Откройте и пересохраните его')
                continue
            table_j_end = 0  # Если больше 10 пустых ячеек - на следующую срочку
            table_k_end = 0  # Если больше 10 пустых строчек - заканчиваем чтение таблицы
            for j, row in enumerate(sheet.rows):
                if table_j_end == 10 and j == 10:
                    break
                snils = 0
                audiofiles = []
                for k, cell in enumerate(row):
                    if cell.value != None:
                        table_j_end = 0
                        table_k_end = 0
                    else:
                        table_j_end += 1
                        table_k_end += 1
                    if table_k_end > 10:
                        break
                    if isSNILS(cell.value):
                        snils = l(cell.value)
                        if snils not in snilsesI:
                            snilsesI.append(snils)
snilsesInput = tuple(snilsesI)
print('\n Уникальных СНИЛС в запросе:', len(snilsesInput))


# Ищем все аудиофайлы isSocium() == True во всех подкаталогах каталога FIND_CATALOG
all_audiofiles = []
for root, dirs, files in os.walk(FIND_CATALOG):
    all_audiofiles += [os.path.join(root, name) for name in files if isSocium(name)]
all_audiofilesExt = {}
for all_audiofile in all_audiofiles:
    rezAudioName = all_audiofile.split('/')[len(all_audiofile.split(('/'))) - 1]
    if rezAudioName.endswith('.wav') or rezAudioName.endswith('.mp3'):
        rezAudioName = rezAudioName[:-4]
    if all_audiofilesExt.get(rezAudioName, None):
        if all_audiofile not in all_audiofilesExt[rezAudioName]:
            all_audiofilesExt[rezAudioName].append(all_audiofile)
    else:
        all_audiofilesExt[rezAudioName] = [all_audiofile]

# Вытаскиваем словарь phonesSNILSES[телефон]=[СНИЛС1,...,СНИЛСn] из Сатурна для Социума
dbconfig_crm = read_config(filename='alone.ini', section='crm')
dbconn = MySQLConnection(**dbconfig_crm)
cursor = dbconn.cursor()
sql = 'SELECT ca.client_phone, cl.`number` FROM saturn_crm.clients AS cl ' \
      'LEFT JOIN saturn_crm.contracts AS co ON co.client_id = cl.client_id ' \
      'LEFT JOIN saturn_crm.callcenter AS ca ON ca.contract_id = co.id ' \
      'WHERE cl.subdomain_id = 6 GROUP BY ca.client_phone'
cursor.execute(sql)
phonesSNILSES = {}
rows = cursor.fetchall()
for row in rows:
    if phonesSNILSES.get(row[0], None):
        if row[1] not in phonesSNILSES[row[0]]:
            phonesSNILSES[row[0]].append(row[1])
    else:
        phonesSNILSES[row[0]] = [row[1]]

# Загружаем найденные из остальных источников
snilsesProblem = {}
snilsesProblemShort = {}
wb = openpyxl.load_workbook(filename=PROBLEMREESTR, read_only=True)
for sheetname in wb.sheetnames:
    sheet = wb[sheetname]
    if not sheet.max_row:
        print('Файл', PROBLEMREESTR, 'Excel некорректно сохранен OpenPyxl. Откройте и пересохраните его')
        continue
    for j, row in enumerate(sheet.rows):
        snils = l(row[0].value)
        snilsProblemAudios = []
        for k, cell in enumerate(row):
            if k and cell.value:
                snilsProblemAudio = isAudio(cell.value)
                if snilsProblemAudio[1] not in snilsProblemAudios:
                    snilsProblemAudios.append(snilsProblemAudio[1])
                if snilsesProblem.get(snils, None):
                    if cell.value not in snilsesProblem[snils]:
                        snilsesProblem[snils].append(cell.value)
                else:
                    snilsesProblem[snils] = [cell.value]
        snilsesProblemShort[snils] = snilsProblemAudios

print('\n Восстановили СНИЛС из остальных источников:', len(snilsesProblem))

# Перебираем файлы и дозаполняем массив [СНИЛС] = [файл1, файл2...]
for all_audiofileExt in all_audiofilesExt:
    all_audiofile = isAudio(all_audiofileExt)
    snils = 0
    if all_audiofile[0] == 'длинный':
        snils = l(all_audiofile[1][20:31])
    elif all_audiofile[0] == 'короткий':
        phone = l(all_audiofile[1][15:26])
        # через телефон найти СНИЛС (если есть)
        if phonesSNILSES.get(phone, None):




# Собираем сначала из точных источников, потом из второстепенных
wb = openpyxl.Workbook(write_only=True)
ws = wb.create_sheet('Остальные')
shure_rez = {}
problem_rez = {}
for snils in snilses:
    if snils_audios_fullpath.get(snils, None):
        shure_rez[snils] = snils_audios_fullpath[snils]
    elif all_snils_audios.get(snils):
        all_snils_audio_vars = []                       # Собираем все файлы сюда
        for all_snils_audio in all_snils_audios:
            if all_audiofilesExt.get(all_snils_audio):
                for all_audiofileExt in all_audiofilesExt[all_snils_audio]:
                    all_snils_audio_vars.append(all_audiofileExt)
        problem_rez[snils] = all_snils_audio_vars
        ws.append([fine_snils(snils)] + all_snils_audio_vars)
wb.save('Остальные.xlsx')
print('Закрыто СНИЛСов из точных источников:', len(shure_rez))
print('Закрыто СНИЛСов из остальных источников:', len(problem_rez))








