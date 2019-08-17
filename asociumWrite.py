import openpyxl, traceback
import os, string, sys, shutil
from collections import Counter
from lib import l, fine_snils_, read_config

FIND_CATALOG = '/media/da3/asteriskBeagleAl/'
#CHANGE_ON_WINDOWS = 'Z:/'
#OUTPUT_CATALOG = 'O:/Документы/Записи/'
OUTPUT_CATALOG = '/media/da3/backup/'
TRUSTREESTR = 'Надежные.xlsx'
PROBLEMREESTR = 'Остальные.xlsx'
REESTRS = '/home/da3/Beagle/потеряшкиАудиозаписи/реестры/'

def isSNILS(snils):
    if snils != None:
        t = str(snils).replace('\n',' ').replace('  ', ' ').replace('  ', ' ').replace('  ', ' ').strip()
        if len(t) > 11:
            if t[3] == '-' and t[7] == '-' and (t[11] == ' ' or t[11] == '_'):
                return True
            else:
                return False
        else:
            return False
    return False

def isAudio(audio):
    if audio != None:
        t = str(audio).replace('\n',' ').replace('  ', ' ').replace('  ', ' ').replace('  ', ' ').strip()
        t1 = t.split('/')[len(t.split(('/'))) - 1]
        if t1.endswith('.'):
            t1 = t1[:-1]
        if t1.endswith('.mp3') or t1.endswith('.wav'):
            t1 = t1[:-4]
        if len(t1) > 26:
            if t1[2] == '.' and t1[5] == '.' and t1[10] == '_' and (t1[13] == '-' or t1[13] == '_') and \
                    (t1[16] == '-' or t1[16] == '_'):
                return ['длинный', t1]
            elif len(''.join([char for i, char in enumerate(t1) if char in string.digits and i < 26])) == 25 \
                    and t1[14] == '_':
                return ['короткий', t1]
            else:
                return ['', audio]
        else:
            return ['', audio]
    return ['', audio]

def isSocium(audio):
    if audio != None:
        t = str(audio).replace('\n',' ').replace('  ', ' ').replace('  ', ' ').replace('  ', ' ').strip()
        t1 = t.split('/')[len(t.split(('/'))) - 1]
        if len(t1) > 26:
            if t1[2] == '.' and t1[5] == '.' and t1[10] == '_' and (t1[13] == '-' or t1[13] == '_') and \
                    (t1[16] == '-' or t1[16] == '_') and (t1[6:10] == '2017' or t1[6:10] == '2018'):
                return True
            elif len(''.join([char for i, char in enumerate(t1) if char in string.digits and i < 26])) == 25 \
                    and t1[14] == '_' and (t1[:4] == '2017' or t1[:4] == '2018'):
                return True
            else:
                return False
        else:
            return False
    return False

# расшифровка любой ошибки
def full_tb_write(*args):
    if not args:
        exc_type, exc_val, exc_tb = sys.exc_info()
        traceback.print_tb(exc_tb, file=sys.stdout)
    elif len(args) == 3:
        exc_type, exc_val, exc_tb = args
        traceback.print_tb(exc_tb, file=sys.stdout)
    elif len(args) == 1:
        exc_type, exc_val, exc_tb = args[0].__class__, args[0], args[0].__traceback__
        traceback.print_tb(exc_tb, file=sys.stdout)

snilsesTrust = {}
snilsesTrustShort = {}
wb = openpyxl.load_workbook(filename=TRUSTREESTR, read_only=True)
for sheetname in wb.sheetnames:
    sheet = wb[sheetname]
    if not sheet.max_row:
        print('Файл', TRUSTREESTR, 'Excel некорректно сохранен OpenPyxl. Откройте и пересохраните его')
        continue
    for j, row in enumerate(sheet.rows):
        snils = l(row[0].value)
        snilsTrustAudios = []
        for k, cell in enumerate(row):
            if k and cell.value:
                snilsTrustAudio = isAudio(cell.value)
                if snilsTrustAudio[1] not in snilsTrustAudios:
                    snilsTrustAudios.append(snilsTrustAudio[1])
                if snilsesTrust.get(snils, None):
                    if cell.value not in snilsesTrust[snils]:
                        snilsesTrust[snils].append(cell.value)
                else:
                    snilsesTrust[snils] = [cell.value]
        snilsesTrustShort[snils] = snilsTrustAudios

for i, snils in enumerate(snilsesTrust):
    sucess = False
    while not sucess:
        try:
            if not os.path.exists(OUTPUT_CATALOG + 'Выгрузки/' + fine_snils_(snils)):
                os.mkdir(OUTPUT_CATALOG + 'Выгрузки/' + fine_snils_(snils))
            audiofilesShort = []
            for audiofile in snilsesTrust[snils]:
                audiofileShort = isAudio(audiofile)[1]
                if os.path.exists(audiofile): #.replace(FIND_CATALOG, CHANGE_ON_WINDOWS)):
                    if audiofileShort in audiofilesShort:
                        if not os.path.exists(OUTPUT_CATALOG + 'Выгрузки/' + fine_snils_(snils) + '/' + audiofileShort +
                                              '-' + str(Counter(audiofilesShort)[audiofileShort]) + audiofile[-4:]):
                            shutil.copy(audiofile #.replace(FIND_CATALOG, CHANGE_ON_WINDOWS)
                                        , OUTPUT_CATALOG + 'Выгрузки/' + fine_snils_(snils) + '/' + audiofileShort +
                                        '-' + str(Counter(audiofilesShort)[audiofileShort]) + audiofile[-4:])
                            audiofilesShort.append(audiofileShort)
                    else:
                        if not os.path.exists(OUTPUT_CATALOG + 'Выгрузки/' + fine_snils_(snils) + '/' +
                                              audiofileShort + audiofile[-4:]):
                            shutil.copy(audiofile #.replace(FIND_CATALOG, CHANGE_ON_WINDOWS)
                                        , OUTPUT_CATALOG + 'Выгрузки/' + fine_snils_(snils) + '/' + audiofileShort +
                                        audiofile[-4:])
                            audiofilesShort.append(audiofileShort)
                else:
                    print('!!! Нет исходного файла', audiofile)
            sucess = True
        except Exception as e:
            full_tb_write(e)
            print('Ошибка - пробуем ещё раз')
    print('Скопировано', i, 'из', len(snilsesTrust))

print('\nТеперь Остальные\n')

snilsesProblem = {}
snilsesProblemShort = {}
wb = openpyxl.load_workbook(filename=PROBLEMREESTR, read_only=True)
for sheetname in wb.sheetnames:
    sheet = wb[sheetname]
    if not sheet.max_row:
        print('Файл', PROBLEMREESTR, 'Excel некорректно сохранен OpenPyxl. Откройте и пересохраните его')
        continue
    for j, row in enumerate(sheet.rows):
        snils = l(row[0].value)
        snilsProblemAudios = []
        for k, cell in enumerate(row):
            if k and cell.value:
                snilsProblemAudio = isAudio(cell.value)
                if snilsProblemAudio[1] not in snilsProblemAudios:
                    snilsProblemAudios.append(snilsProblemAudio[1])
                if snilsesProblem.get(snils, None):
                    if cell.value not in snilsesProblem[snils]:
                        snilsesProblem[snils].append(cell.value)
                else:
                    snilsesProblem[snils] = [cell.value]
        snilsesProblemShort[snils] = snilsProblemAudios

for i, snils in enumerate(snilsesProblem):
    sucess = False
    while not sucess:
        try:
            if not os.path.exists(OUTPUT_CATALOG + 'Остальные/' + fine_snils_(snils)):
                os.mkdir(OUTPUT_CATALOG + 'Остальные/' + fine_snils_(snils))
            audiofilesShort = []
            for audiofile in snilsesProblem[snils]:
                audiofileShort = isAudio(audiofile)[1]
                if os.path.exists(audiofile): #.replace(FIND_CATALOG, CHANGE_ON_WINDOWS)):
                    if audiofileShort in audiofilesShort:
                        if not os.path.exists(OUTPUT_CATALOG + 'Остальные/' + fine_snils_(snils) + '/' + audiofileShort +
                                              '-' + str(Counter(audiofilesShort)[audiofileShort]) + audiofile[-4:]):
                            shutil.copy(audiofile #.replace(FIND_CATALOG, CHANGE_ON_WINDOWS)
                                        , OUTPUT_CATALOG + 'Остальные/' + fine_snils_(snils) + '/' + audiofileShort +
                                        '-' + str(Counter(audiofilesShort)[audiofileShort]) + audiofile[-4:])
                            audiofilesShort.append(audiofileShort)
                    else:
                        if not os.path.exists(OUTPUT_CATALOG + 'Остальные/' + fine_snils_(snils) + '/' +
                                              audiofileShort + audiofile[-4:]):
                            shutil.copy(audiofile #.replace(FIND_CATALOG, CHANGE_ON_WINDOWS)
                                        , OUTPUT_CATALOG + 'Остальные/' + fine_snils_(snils) + '/' + audiofileShort +
                                        audiofile[-4:])
                            audiofilesShort.append(audiofileShort)
                else:
                    print('!!! Нет исходного файла', audiofile)
            sucess = True
        except Exception as e:
            full_tb_write(e)
            print('Ошибка - пробуем ещё раз')
    print('Скопировано', i, 'из', len(snilsesProblem))







